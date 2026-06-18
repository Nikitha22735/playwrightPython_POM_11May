
import csv
import os

from dotenv import load_dotenv
from openpyxl import load_workbook
import pytest
import json

from utils.jsonHandling import jsonData



def test_cli():
    print(os.getenv("us1"))
    print(os.getenv("pw1"))

#cmd
# set us1=username1&&set pw1=password1&&pytest -m datathandling --headed -s
#powershell
# $env us1=username1;$env pw1=password1;pytest -m datathandling --headed -s   

#pip install python-dotenv
# @pytest.mark.datathandling
def test_cli_env():
    load_dotenv(os.getenv("file"))
    print(os.getenv("us2"))
    print(os.getenv("pw2"))


# ========================================json=================================
# @pytest.mark.datathandling
def test_json():
    # data = open("testData\\data1.json")
    with open("testData\\data1.json") as data:
        formattedData = json.load(data)
        print(formattedData["positive"]["username"])
    # data.close()



# @pytest.mark.datathandling
def test_json_1():
    testData = jsonData("testData\\data1.json")
    print(testData)
    

# ===========================================csv====================================
# @pytest.mark.datathandling
def test_read_csv():
    with open("testData\\credentails.csv") as data:
        formattedData = csv.DictReader(data)
        values = []
        for i in formattedData:
            values.append(i)

    print(values[1]["username"])


# @pytest.mark.datathandling
def test_write_csv():
    with open("testData\\credentails.csv", mode='w', newline="") as data:
        formattedData = csv.DictWriter(data, fieldnames=["username","password"])
        formattedData.writeheader()
        formattedData.writerow({'username': 'tripur123_4', 'password': 'welcome123'})


# ================================================excel===================

        
#pip install openpyxl
# @pytest.mark.datathandling
def test_read_excel():
    workbook = load_workbook("testData/sample_creds.xlsx")
    sheet = workbook["Sheet2"]
    values = []
    for i in sheet.iter_cols(min_col=1, values_only=True):
        values.append(i)
    print(values)


# @pytest.mark.datathandling
def test_write_excel():
    workbook = load_workbook("testData/sample_creds.xlsx")
    sheet = workbook.create_sheet("Sheet5")
    # sheet = workbook["sheet5"]
    # sheet.append(["test_newline","line2"])
    sheet["A2"]="test_1234"
    # sheet.delete_rows(3,1)
    workbook.save("testData/sample_creds.xlsx")


@pytest.mark.datathandling
@pytest.mark.parametrize("a,b",[(2,3),(4,5)])
def test_sum(a,b):
    print(a),
    print(b)
    c=10
    assert c>(a+b)






